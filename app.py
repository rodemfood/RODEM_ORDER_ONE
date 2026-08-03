import io
import json
import os
import re
import secrets
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, make_response, redirect, render_template, request, send_file
from werkzeug.middleware.proxy_fix import ProxyFix
from sqlalchemy import Column, Integer, String, Text, create_engine, select, update
from sqlalchemy.orm import declarative_base, sessionmaker
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent

# 운영 환경에서는 Render의 DATABASE_URL(PostgreSQL)을 사용합니다.
# DATABASE_URL이 아직 연결되지 않은 경우에도 배포가 실패하지 않도록
# 테스트 전용 SQLite를 /tmp에 생성합니다. /tmp 데이터는 재시작 시 사라집니다.
RAW_DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USING_POSTGRES = False
DATABASE_FALLBACK_REASON = ""

def normalize_database_url(raw_url):
    url = (raw_url or "").strip()
    if url.startswith("postgres://"):
        return url.replace("postgres://", "postgresql+psycopg://", 1)
    if url.startswith("postgresql://") and "+psycopg" not in url:
        return url.replace("postgresql://", "postgresql+psycopg://", 1)
    return url

def build_engine():
    global USING_POSTGRES, DATABASE_FALLBACK_REASON
    candidates = []
    normalized = normalize_database_url(RAW_DATABASE_URL)
    if normalized:
        candidates.append((normalized, True))
    candidates.append(("sqlite:////tmp/rodem_order_one.db", False))

    last_error = None
    for url, is_postgres in candidates:
        try:
            kwargs = {"pool_pre_ping": True}
            if url.startswith("sqlite"):
                kwargs["connect_args"] = {"check_same_thread": False, "timeout": 30}
            else:
                kwargs.update({"pool_recycle": 300, "pool_size": 5, "max_overflow": 5})
            candidate = create_engine(url, **kwargs)
            with candidate.connect() as conn:
                conn.exec_driver_sql("SELECT 1")
            USING_POSTGRES = is_postgres
            if last_error:
                DATABASE_FALLBACK_REASON = clean_text(str(last_error), 300) if "clean_text" in globals() else str(last_error)[:300]
            return candidate
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"데이터베이스를 초기화할 수 없습니다: {last_error}")

engine = build_engine()
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
Base = declarative_base()

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
app.config["JSON_AS_ASCII"] = False
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024


class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True)
    token = Column(String(128), unique=True, nullable=False, index=True)
    company = Column(String(100), nullable=False)
    receiver = Column(String(50), nullable=False)
    phone = Column(String(30), nullable=False)
    postal_code = Column(String(20), default="")
    address = Column(String(250), nullable=False)
    created_at = Column(String(19), nullable=False)


class Order(Base):
    __tablename__ = "orders"
    id = Column(Integer, primary_key=True)
    order_no = Column(String(40), unique=True, nullable=False, index=True)
    request_id = Column(String(80), unique=True, nullable=False, index=True)
    customer_token = Column(String(128), nullable=False, index=True)
    company = Column(String(100), nullable=False)
    receiver = Column(String(50), nullable=False)
    phone = Column(String(30), nullable=False)
    postal_code = Column(String(20), default="")
    address = Column(String(250), nullable=False)
    memo = Column(String(500), default="")
    items_json = Column(Text, nullable=False)
    total_qty = Column(Integer, nullable=False)
    status = Column(String(20), nullable=False, default="신규")
    created_at = Column(String(19), nullable=False)
    invoiced_at = Column(String(19), default="")


try:
    Base.metadata.create_all(engine)
except Exception as exc:
    # 외부 DB가 일시적으로 실패해도 Render 배포 자체가 중단되지 않도록 /tmp SQLite로 자동 전환합니다.
    DATABASE_FALLBACK_REASON = str(exc)[:300]
    USING_POSTGRES = False
    engine.dispose()
    engine = create_engine(
        "sqlite:////tmp/rodem_order_one.db",
        pool_pre_ping=True,
        connect_args={"check_same_thread": False, "timeout": 30},
    )
    SessionLocal.configure(bind=engine)
    Base.metadata.create_all(engine)


def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def make_order_no():
    return "R" + datetime.now().strftime("%y%m%d%H%M%S%f")[:18]


def clean_text(value, max_len=200):
    return re.sub(r"[\x00-\x1f\x7f]", " ", str(value or "")).strip()[:max_len]


def clean_phone(value):
    return clean_text(value, 30)


def customer_to_dict(row):
    return {
        "company": row.company,
        "receiver": row.receiver,
        "phone": row.phone,
        "postal_code": row.postal_code or "",
        "address": row.address,
    }


def order_to_dict(row):
    return {
        "id": row.id,
        "order_no": row.order_no,
        "company": row.company,
        "receiver": row.receiver,
        "phone": row.phone,
        "postal_code": row.postal_code or "",
        "address": row.address,
        "memo": row.memo or "",
        "items": json.loads(row.items_json),
        "total_qty": row.total_qty,
        "status": row.status,
        "created_at": row.created_at,
        "invoiced_at": row.invoiced_at or "",
    }


def get_customer():
    token = request.cookies.get("rodem_customer", "")
    if not token:
        return None
    with SessionLocal() as session:
        return session.scalar(select(Customer).where(Customer.token == token))


def logen_product_text(items):
    # 로젠 필수 규칙: 첫 제품부터 모든 제품 앞에 # 삽입
    return "".join(f"#{clean_text(item['name'], 100).replace('#', '')}{int(item['qty'])}" for item in items)


@app.get("/")
def home():
    return redirect("/order")


@app.get("/health")
def health():
    return jsonify(
        ok=True,
        service="RODEM ORDER ONE",
        database="postgresql" if USING_POSTGRES else "temporary-sqlite",
        persistent=USING_POSTGRES,
        time=now_text(),
        fallback_reason=DATABASE_FALLBACK_REASON,
    )


@app.get("/order")
def order_page():
    return render_template("customer.html")


@app.get("/staff")
def staff_page():
    return render_template("staff.html")


@app.get("/api/customer/me")
def customer_me():
    customer = get_customer()
    return jsonify(customer=customer_to_dict(customer) if customer else None)


@app.post("/api/customer/register")
def register_customer():
    data = request.get_json(silent=True) or {}
    values = {
        "company": clean_text(data.get("company"), 100),
        "receiver": clean_text(data.get("receiver"), 50),
        "phone": clean_phone(data.get("phone")),
        "postal_code": clean_text(data.get("postal_code"), 20),
        "address": clean_text(data.get("address"), 250),
    }
    if not values["company"] or not values["receiver"] or not values["phone"] or not values["address"]:
        return jsonify(error="업체명, 받는 분, 연락처, 배송지 주소를 입력해 주세요."), 400

    token = secrets.token_urlsafe(32)
    customer = Customer(token=token, created_at=now_text(), **values)
    with SessionLocal() as session:
        session.add(customer)
        session.commit()

    response = make_response(jsonify(customer=customer_to_dict(customer)))
    response.set_cookie(
        "rodem_customer", token, max_age=31536000 * 3,
        httponly=True, samesite="Lax", secure=request.is_secure,
    )
    return response




@app.put("/api/customer/me")
def update_customer():
    customer = get_customer()
    if not customer:
        return jsonify(error="저장된 고객 정보를 찾을 수 없습니다."), 401

    data = request.get_json(silent=True) or {}
    values = {
        "company": clean_text(data.get("company"), 100),
        "receiver": clean_text(data.get("receiver"), 50),
        "phone": clean_phone(data.get("phone")),
        "postal_code": clean_text(data.get("postal_code"), 20),
        "address": clean_text(data.get("address"), 250),
    }
    if not values["company"] or not values["receiver"] or not values["phone"] or not values["address"]:
        return jsonify(error="업체명, 받는 분, 연락처, 배송지 주소를 입력해 주세요."), 400

    with SessionLocal() as session:
        row = session.scalar(select(Customer).where(Customer.token == customer.token))
        if not row:
            return jsonify(error="저장된 고객 정보를 찾을 수 없습니다."), 404
        for key, value in values.items():
            setattr(row, key, value)
        session.commit()
        session.refresh(row)
        return jsonify(customer=customer_to_dict(row))


@app.get("/api/customer/orders")
def customer_orders():
    customer = get_customer()
    if not customer:
        return jsonify(error="저장된 고객 정보를 찾을 수 없습니다."), 401
    with SessionLocal() as session:
        rows = session.scalars(
            select(Order)
            .where(Order.customer_token == customer.token)
            .order_by(Order.id.desc())
            .limit(20)
        ).all()
        return jsonify(orders=[order_to_dict(row) for row in rows])


@app.post("/api/orders")
def create_order():
    customer = get_customer()
    if not customer:
        return jsonify(error="배송지 정보를 먼저 저장해 주세요."), 401

    data = request.get_json(silent=True) or {}
    request_id = clean_text(data.get("request_id"), 80) or secrets.token_urlsafe(24)
    items = []
    for item in (data.get("items") or [])[:8]:
        name = clean_text(item.get("name"), 100).replace("#", "")
        try:
            qty = int(item.get("qty", 0))
        except (TypeError, ValueError):
            qty = 0
        if name and 0 < qty <= 999999:
            items.append({"name": name, "qty": qty})
    if not items:
        return jsonify(error="제품명과 낱개 수량을 한 개 이상 입력해 주세요."), 400

    with SessionLocal() as session:
        existing = session.scalar(select(Order).where(Order.request_id == request_id))
        if existing:
            return jsonify(order_no=existing.order_no, total_qty=existing.total_qty, duplicate=True)

        total_qty = sum(item["qty"] for item in items)
        row = Order(
            order_no=make_order_no(), request_id=request_id,
            customer_token=customer.token, company=customer.company,
            receiver=customer.receiver, phone=customer.phone,
            postal_code=customer.postal_code or "", address=customer.address,
            memo=clean_text(data.get("memo"), 500),
            items_json=json.dumps(items, ensure_ascii=False), total_qty=total_qty,
            status="신규", created_at=now_text(), invoiced_at="",
        )
        session.add(row)
        session.commit()
        return jsonify(order_no=row.order_no, total_qty=row.total_qty, duplicate=False)


@app.get("/api/staff/orders")
def staff_orders():
    with SessionLocal() as session:
        rows = session.scalars(select(Order).order_by(Order.id.desc())).all()
        return jsonify(orders=[order_to_dict(row) for row in rows])


@app.post("/api/staff/mark-new")
def mark_new():
    data = request.get_json(silent=True) or {}
    try:
        order_id = int(data.get("id"))
    except (TypeError, ValueError):
        return jsonify(error="주문번호가 올바르지 않습니다."), 400
    with SessionLocal() as session:
        session.execute(update(Order).where(Order.id == order_id).values(status="신규", invoiced_at=""))
        session.commit()
    return jsonify(ok=True)


def selected_orders(ids):
    with SessionLocal() as session:
        return session.scalars(select(Order).where(Order.id.in_(ids)).order_by(Order.id)).all()


@app.post("/api/staff/export-logen")
def export_logen():
    data = request.get_json(silent=True) or {}
    try:
        ids = sorted({int(v) for v in data.get("ids", [])})
    except (TypeError, ValueError):
        ids = []
    if not ids:
        return jsonify(error="송장을 생성할 주문을 선택해 주세요."), 400

    rows = selected_orders(ids)
    if not rows:
        return jsonify(error="선택한 주문을 찾을 수 없습니다."), 404

    # 로젠 공식 열 순서에 맞춘 .xls 출력
    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Sheet4")
    headers = [
        "수화주명", "전화번호", "휴대폰", "우편번호", "주소", "상품명",
        "박스수량", "배송메시지", "선착분", "운임", "", "주문자명"
    ]
    header_style = xlwt.easyxf(
        "font: bold on, color white; pattern: pattern solid, fore_colour green;"
        "align: horiz center, vert center; borders: left thin, right thin, top thin, bottom thin;"
    )
    cell_style = xlwt.easyxf(
        "align: vert top, wrap on; borders: left thin, right thin, top thin, bottom thin;"
    )
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)
    widths = [18, 18, 18, 12, 50, 70, 10, 40, 10, 10, 5, 22]
    for i, width in enumerate(widths):
        ws.col(i).width = width * 256

    for r, row in enumerate(rows, start=1):
        items = json.loads(row.items_json)
        values = [
            row.receiver,
            row.phone,
            row.phone,
            row.postal_code or "",
            row.address,
            logen_product_text(items),
            1,
            row.memo or "",
            "",
            "",
            "",
            row.company,
        ]
        for c, value in enumerate(values):
            ws.write(r, c, value, cell_style)

    with SessionLocal() as session:
        session.execute(
            update(Order).where(Order.id.in_(ids)).values(status="송장생성", invoiced_at=now_text())
        )
        session.commit()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output, as_attachment=True,
        download_name=f"RODEM_LOGEN_{datetime.now():%Y%m%d_%H%M}.xls",
        mimetype="application/vnd.ms-excel",
    )


def style_xlsx_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="188754")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


@app.post("/api/staff/export-backup")
def export_backup():
    with SessionLocal() as session:
        rows = session.scalars(select(Order).order_by(Order.id.desc())).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "주문백업"
    style_xlsx_header(ws, [
        "상태", "주문번호", "접수시간", "업체명", "받는 분", "연락처",
        "우편번호", "주소", "주문상품", "로젠상품명", "총 낱개수량",
        "배송요청사항", "송장생성시간"
    ])
    for row in rows:
        items = json.loads(row.items_json)
        products = " / ".join(f"{i['name']} {i['qty']}개" for i in items)
        ws.append([
            row.status, row.order_no, row.created_at, row.company, row.receiver,
            row.phone, row.postal_code or "", row.address, products,
            logen_product_text(items), row.total_qty, row.memo or "", row.invoiced_at or ""
        ])
    for col, width in {
        "A":12,"B":24,"C":20,"D":22,"E":16,"F":18,"G":12,"H":45,
        "I":65,"J":65,"K":14,"L":38,"M":20
    }.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output, as_attachment=True,
        download_name=f"RODEM_ORDER_BACKUP_{datetime.now():%Y%m%d_%H%M}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# 거래처가 자체 양식으로 보낸 엑셀 발주서를 로젠 업로드 형식으로 변환합니다.
HEADER_ALIASES = {
    "receiver": ["수화주명", "받는분", "받는 분", "수령인", "수취인", "배송처", "점포명", "매장명"],
    "mobile": ["휴대폰", "휴대전화", "핸드폰", "연락처", "휴대폰번호"],
    "phone": ["전화", "전화번호", "일반전화"],
    "postal": ["우편번호", "우편 번호", "우편"],
    "address": ["주소", "배송지", "배송주소", "배송지주소"],
    "product": ["물품명", "상품명", "제품명", "품목명", "주문상품"],
    "qty": ["주문수량", "총수량", "수량", "주문 수량"],
    "company": ["주문자명", "업체명", "거래처명", "발주처"],
    "memo": ["배송메시지", "배송 요청사항", "요청사항", "비고"],
}


def normalized_header(value):
    return re.sub(r"[\s_\-./()]", "", clean_text(value, 80)).lower()


def find_header_map(ws):
    aliases = {key: {normalized_header(v) for v in values} for key, values in HEADER_ALIASES.items()}
    best = None
    for row_idx in range(1, min(ws.max_row, 25) + 1):
        found = {}
        for col_idx in range(1, min(ws.max_column, 80) + 1):
            value = ws.cell(row_idx, col_idx).value
            key_text = normalized_header(value)
            if not key_text:
                continue
            for key, names in aliases.items():
                if key_text in names and key not in found:
                    found[key] = col_idx
        score = sum(k in found for k in ("receiver", "address", "product", "qty"))
        if score >= 3 and (best is None or score > best[0]):
            best = (score, row_idx, found)
    return (best[1], best[2]) if best else (None, {})


def excel_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value, 500)


def excel_qty(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return int(float(match.group())) if match else 0


def clean_import_product(value):
    name = excel_text(value).replace("#", "").strip()
    # 거래처 발주서의 포장규격 표기는 제거하고 실제 상품명만 송장에 표시합니다.
    name = re.sub(r"\s+\d+(?:\.\d+)?\s*(?:g|kg|ml|l)\s*\d+\s*개\s*$", "", name, flags=re.I)
    name = re.sub(r"\s+\d+\s*입\s*$", "", name)
    return name.strip()


def load_excel_workbook(upload):
    filename = (upload.filename or "").lower()
    raw = upload.read()
    if not raw:
        raise ValueError("빈 파일입니다.")
    if filename.endswith((".xlsx", ".xlsm")):
        return load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
    if filename.endswith(".xls"):
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("구형 .xls 파일을 읽는 구성요소가 없습니다.") from exc
        book = xlrd.open_workbook(file_contents=raw)
        # openpyxl과 동일한 최소 인터페이스로 변환합니다.
        temp = Workbook()
        temp.remove(temp.active)
        for sheet in book.sheets():
            ws = temp.create_sheet(sheet.name[:31])
            for r in range(sheet.nrows):
                for c in range(sheet.ncols):
                    ws.cell(r + 1, c + 1, sheet.cell_value(r, c))
        return temp
    raise ValueError(".xlsx, .xlsm 또는 .xls 파일만 업로드할 수 있습니다.")


def parse_order_workbook(upload):
    wb = load_excel_workbook(upload)
    grouped = {}
    warnings = []
    recognized_sheets = []
    image_sheets = []

    for ws in wb.worksheets:
        image_count = len(getattr(ws, "_images", []))
        if image_count:
            image_sheets.append(f"{ws.title}({image_count}개 이미지)")
        header_row, columns = find_header_map(ws)
        if not header_row:
            if image_count:
                continue
            nonempty = sum(1 for row in ws.iter_rows(values_only=True) for v in row if v not in (None, ""))
            if nonempty:
                warnings.append(f"'{ws.title}' 탭은 표 머리글을 인식하지 못해 제외했습니다.")
            continue

        recognized_sheets.append(ws.title)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            def val(key):
                col = columns.get(key)
                return ws.cell(row_idx, col).value if col else None

            product = clean_import_product(val("product"))
            qty = excel_qty(val("qty"))
            receiver = excel_text(val("receiver"))
            if product in ("합계", "총계", "소계") or not product or qty <= 0:
                continue

            mobile = excel_text(val("mobile"))
            phone = excel_text(val("phone"))
            contact = mobile or phone
            postal = excel_text(val("postal"))
            address = excel_text(val("address"))
            company = excel_text(val("company")) or ws.title
            memo = excel_text(val("memo"))

            # 같은 수화주의 여러 상품행은 로젠 송장 한 줄로 묶습니다.
            key = (ws.title, receiver, contact, address, postal, company, memo)
            if key not in grouped:
                grouped[key] = {
                    "source_sheet": ws.title,
                    "company": company,
                    "receiver": receiver,
                    "phone": contact,
                    "postal_code": postal,
                    "address": address,
                    "memo": memo,
                    "items": [],
                }
            grouped[key]["items"].append({"name": product, "qty": qty})

    rows = []
    for index, row in enumerate(grouped.values(), start=1):
        errors = []
        notices = []
        if not row["receiver"]:
            errors.append("받는 분 누락")
        if not row["phone"]:
            errors.append("연락처 누락")
        if not row["address"]:
            errors.append("주소 누락")
        if not row["postal_code"]:
            notices.append("우편번호 없음")
        row["row_id"] = index
        row["total_qty"] = sum(int(i["qty"]) for i in row["items"])
        row["logen_product"] = logen_product_text(row["items"])
        row["valid"] = not errors
        row["errors"] = errors
        row["notices"] = notices
        rows.append(row)

    if image_sheets:
        warnings.append("이미지로만 구성된 탭은 자동 변환하지 않았습니다: " + ", ".join(image_sheets))
    if not rows:
        raise ValueError("자동 변환할 수 있는 주문 데이터를 찾지 못했습니다.")
    return {
        "rows": rows,
        "recognized_sheets": recognized_sheets,
        "warnings": warnings,
        "summary": {
            "total": len(rows),
            "valid": sum(1 for r in rows if r["valid"]),
            "error": sum(1 for r in rows if not r["valid"]),
            "total_qty": sum(r["total_qty"] for r in rows if r["valid"]),
        },
    }


@app.post("/api/staff/import-excel")
def import_partner_excel():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify(error="변환할 엑셀 파일을 선택해 주세요."), 400
    try:
        result = parse_order_workbook(upload)
        return jsonify(result)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    except Exception as exc:
        app.logger.exception("거래처 발주서 변환 실패")
        return jsonify(error=f"엑셀 파일을 읽는 중 오류가 발생했습니다: {clean_text(exc, 200)}"), 400


@app.post("/api/staff/export-imported-logen")
def export_imported_logen():
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    selected = []
    for row in rows[:5000]:
        if not row.get("selected", True) or not row.get("valid", False):
            continue
        items = []
        for item in (row.get("items") or [])[:100]:
            name = clean_import_product(item.get("name"))
            qty = excel_qty(item.get("qty"))
            if name and qty > 0:
                items.append({"name": name, "qty": qty})
        receiver = clean_text(row.get("receiver"), 100)
        phone = clean_phone(row.get("phone"))
        address = clean_text(row.get("address"), 300)
        if not receiver or not phone or not address or not items:
            continue
        selected.append({
            "receiver": receiver,
            "phone": phone,
            "postal_code": clean_text(row.get("postal_code"), 20),
            "address": address,
            "company": clean_text(row.get("company"), 100),
            "memo": clean_text(row.get("memo"), 500),
            "items": items,
        })
    if not selected:
        return jsonify(error="정상 변환된 주문을 한 건 이상 선택해 주세요."), 400

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Sheet4")
    headers = [
        "수화주명", "전화번호", "휴대폰", "우편번호", "주소", "상품명",
        "박스수량", "배송메시지", "선착분", "운임", "", "주문자명"
    ]
    header_style = xlwt.easyxf(
        "font: bold on, color white; pattern: pattern solid, fore_colour green;"
        "align: horiz center, vert center; borders: left thin, right thin, top thin, bottom thin;"
    )
    cell_style = xlwt.easyxf(
        "align: vert top, wrap on; borders: left thin, right thin, top thin, bottom thin;"
    )
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)
    widths = [18, 18, 18, 12, 50, 70, 10, 40, 10, 10, 5, 22]
    for i, width in enumerate(widths):
        ws.col(i).width = width * 256
    for r, row in enumerate(selected, start=1):
        values = [
            row["receiver"], row["phone"], row["phone"], row["postal_code"],
            row["address"], logen_product_text(row["items"]), 1, row["memo"],
            "", "", "", row["company"],
        ]
        for c, value in enumerate(values):
            ws.write(r, c, value, cell_style)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output, as_attachment=True,
        download_name=f"RODEM_PARTNER_LOGEN_{datetime.now():%Y%m%d_%H%M}.xls",
        mimetype="application/vnd.ms-excel",
    )


@app.errorhandler(413)
def too_large(_):
    return jsonify(error="입력 데이터가 너무 큽니다."), 413


if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get("PORT", "5000"))
    serve(app, host="0.0.0.0", port=port)
